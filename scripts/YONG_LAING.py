import openpyxl
from openpyxl.utils import get_column_letter
import xlrd
import os

def convert_xls_to_xlsx(xls_filename, xlsx_filename):
    """
    Converts an XLS file to XLSX format.
    Args:
        xls_filename (str): The path to the XLS file.
        xlsx_filename (str): The path where the XLSX file will be saved.
    """
    try:
        workbook = xlrd.open_workbook(xls_filename)
        wb = openpyxl.Workbook()
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            new_sheet = wb.create_sheet(title=sheet_name)
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    new_sheet.cell(row=row + 1, column=col + 1, value=cell_value)
        # Remove the default sheet if it's still present
        if "Sheet" in wb.sheetnames:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)
        wb.save(xlsx_filename)
        print(f"Successfully converted '{xls_filename}' to '{xlsx_filename}'")
    except Exception as e:
        print(f"Error converting XLS to XLSX: {e}")
        return False
    return True

def organize_data(filename):
    """
    Organizes data from 'INVOICE' sheet in an Excel file
    and writes the organized data to a new Excel file.

    Args:
    filename (str): The path to the Excel file containing 'INVOICE' sheet.
    """
    try:
        # Load the input workbook and select the sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['INVOICE']

        # Create a new workbook for the output
        output_wb = openpyxl.Workbook()
        organized_sheet = output_wb.active
        organized_sheet.title = 'Organized'

        # Define the headers and their indexes
        headers = ['PO Number ', 'DESCRIPTION OF GOODS  ', 'PARTS  NO.', 'w/o', 'Longsys SKU#', 'Brand', 'Q\'ty (pcs)', 'Unit Price', 'Amount (USD)']
        headers_idx = {header: -1 for i, header in enumerate(headers)}

        # Initialize variables
        made_in_taiwan = False
        made_in_china = False
        i = 0
        j = 1

        # Loop through the rows in column A using index i
        while i <= sheet.max_row:
            i += 1
            col_1 = sheet.cell(row=i, column=1).value
            col_2 = sheet.cell(row=i, column=2).value

            # Check if the row is within the "Made In Taiwan" section
            if made_in_taiwan:
                # Check if the row is within the "Made In China" section
                if made_in_china:
                    if not col_2:
                        break
                elif col_1 == 'Made In China' or col_2 == 'Made In China':
                    made_in_china = True
                    continue
                elif not col_2:
                    continue

            # Check if the row contains headers
            elif col_1 in headers:

                # Find the column indexes for the headers
                for col_idx in range(1, sheet.max_column + 1):
                    col_val = sheet.cell(row=i, column=col_idx).value
                    if col_val in headers:
                        headers_idx[col_val] = col_idx
                
                # Remove headers with index -1
                headers = [header for header in headers if headers_idx[header] > 0]
                headers_idx = {headers[i]: i for i in range(len(headers))}

                made_in_taiwan = True
                i += 1
                continue

            else:
                continue

            # Extract the row data based on the headers
            row = [sheet.cell(row=i, column=headers_idx[header]+1).value for header in headers]

            k = 1
            for header in headers:
                if header in ['PO Number ', 'DESCRIPTION OF GOODS  ', 'PARTS  NO.', 'w/o', 'Longsys SKU#']:
                    input_value = row[k - 1]
                    if input_value:
                        if header == 'PO Number ':
                            input_value = f'PO Number {input_value}'
                        elif header == 'PARTS  NO.':
                            input_value = f'PARTS NO. {input_value}'
                    else:
                        input_value = ''
                    organized_sheet.cell(row=j, column=k, value=input_value)
                    k += 1
                else:
                    break
            
            # Add quantity to the organized sheet
            qty = row[headers_idx['Q\'ty (pcs)']]
            qty_cell = organized_sheet.cell(row=j, column=k, value=qty)
            if isinstance(qty, (int, float)):
                qty_cell.number_format = '#,##0'
            k += 1

            # Add "PCS" to the organized sheet
            organized_sheet.cell(row=j, column=k, value='PCS')
            k += 1

            # Add unit price to the organized sheet
            unit_price = row[headers_idx['Unit Price']]
            unit_price_cell = organized_sheet.cell(row=j, column=k, value=unit_price)
            if isinstance(unit_price, (int, float)):
                unit_price_cell.number_format = '#,##0.0000'
            k += 1

            # Add brand to the organized sheet
            brand = row[headers_idx['Brand']]
            if brand:
                organized_sheet.cell(row=j, column=k, value=brand)
            else:
                organized_sheet.cell(row=j, column=k, value='NO BRAND')
            k += 1

            # Add fixed value to the organized sheet
            organized_sheet.cell(row=j, column=k, value='8523.51.00.00-0')
            k += 1

            # Add country code to the organized sheet
            if made_in_china:
                organized_sheet.cell(row=j, column=k, value='8B')
            elif made_in_taiwan:
                organized_sheet.cell(row=j, column=k, value='06')

            # Adjust column widths
            for col in range(1, k + 1):
                max_length = 0
                column = get_column_letter(col)
                for cell in organized_sheet[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                organized_sheet.column_dimensions[column].width = adjusted_width

            j += 1

        # Save the organized data to a new file in the 'uploads' folder
        output_filename = os.path.join("uploads", "Organized_Data.xlsx")
        output_wb.save(output_filename)
        print(f"Organized data saved to '{output_filename}'")

    except FileNotFoundError:
        print(f"Error: File not found at '{filename}'.")
    except Exception as e:
        print(f"An error occurred: {e}")
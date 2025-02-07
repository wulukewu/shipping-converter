import openpyxl

def read_xlsx_and_output_txt(xlsx_path, txt_path):
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    # Open the text file for writing
    with open(txt_path, 'w', encoding='utf-8') as txt_file:

        # Write initial content to the text file
        content = f'其他申報事項:\nInvoice No.{"*"*20}'
        txt_file.write(content + '\n')

        # Define headers and initialize their indexes
        headers = ['進口日期', '報單號碼', '項次', '容量', '進料單價(USD)', '詠聯組裝費(USD)', '核銷報單項次', '核銷數量']
        headers_idx = {header: -1 for idx, header in enumerate(headers)}

        # Initialize variables
        i = 1

        date = None
        num = None
        storage = None
        unit_price = None
        assembly_fee = None
        verification_item = None
        verification_quantity = None

        first_print_date = True

        # Find the column indexes for the headers
        for col_idx in range(1, sheet.max_column + 1):
            col_val = sheet.cell(row=i, column=col_idx).value
            if col_val in headers:
                headers_idx[col_val] = col_idx

        # Remove headers with index -1
        headers = [header for header in headers if headers_idx[header] > 0]
        headers_idx = {header: headers_idx[header] for header in headers}
        # print(headers_idx)

        # Iterate through each row in the sheet
        while i <= sheet.max_row:
            i += 1

            # Read the row data based on header indexes
            row = [sheet.cell(row=i, column=headers_idx[header]).value for header in headers]
            row_data = dict(zip(headers, row))
            # print(row_data)

            # Extract values from the row data
            date_tmp = row_data['進口日期']
            num_tmp = row_data['報單號碼']
            item = row_data['項次']
            storage_tmp = row_data['容量']
            unit_price_tmp = row_data['進料單價(USD)']
            assembly_fee_tmp = row_data['詠聯組裝費(USD)']
            verification_item_tmp = row_data['核銷報單項次']
            verification_quantity_tmp = row_data['核銷數量']

            # Write verification item and quantity if they exist
            if verification_item is not None and verification_quantity is not None:
                content = f'核銷報單項次:{verification_item} 核銷數量:{verification_quantity}'
                txt_file.write(content + '\n')
                verification_item = None
                verification_quantity = None

            # Check if the row is empty
            if len(list(filter(None, row))) == 0:
                break
            # Check if the row has only one non-empty cell
            elif len(list(filter(None, row))) == 1:
                content = ''.join([str(val) for val in list(filter(None, row))])
                txt_file.write(content + '\n')
                continue
            # Check if the item is None
            elif item is None:
                if date_tmp is not None and num_tmp is not None:
                    content = f'進口日期:{date} 報單號碼:{num}'
                    txt_file.write(content + '\n')
                    content = f'進口日期:{date_tmp} 報單號碼:{num_tmp}'
                    txt_file.write(content + '\n')
                    date = None
                    num = None
                    # i += 1
                continue
            else:
                if date_tmp is not None and num_tmp is not None:
                    if date is None and num is None:
                        pass
                    elif first_print_date:
                        first_print_date = False
                        date = date_tmp
                        num = num_tmp
                    else:
                        content = f'進口日期:{date} 報單號碼:{num}'
                        txt_file.write(content + '\n')

                    date = date_tmp
                    num = num_tmp

            # Update storage, unit price, assembly fee, and verification quantity if they exist
            if storage_tmp is not None:
                storage = storage_tmp
            if unit_price_tmp is not None:
                unit_price = unit_price_tmp
            if assembly_fee_tmp is not None:
                assembly_fee = assembly_fee_tmp
            if verification_quantity_tmp is not None:
                verification_quantity_tmp = verification_quantity_tmp.strip()

            # Format unit price if it contains '/pc'
            if isinstance(unit_price, str) and '/pc' in unit_price:
                try:
                    number, unit = unit_price.split('/')
                    unit_price = f'{float(number):.4f}/{unit}'
                except ValueError:
                    pass
            elif isinstance(float(unit_price), (int, float)):
                unit_price = f'{float(unit_price):.4f}'
            
            # Format assembly fee if it contains '/set'
            if isinstance(assembly_fee, str) and '/set' in assembly_fee:
                try:
                    number, unit = assembly_fee.split('/')
                    assembly_fee = f'{float(number):.4f}/{unit}'
                except ValueError:
                    pass
            elif isinstance(float(assembly_fee), (int, float)):
                assembly_fee = f'{float(assembly_fee):.4f}'

            # Format verification quantity
            if isinstance(verification_quantity_tmp, (int, float)):
                verification_quantity_tmp = f"{verification_quantity_tmp:,}"

            # Write item details to the text file
            content = f'項次:{item} 容量:{storage}\n進料單價(USD):{unit_price} 詠聯組裝費(USD):{assembly_fee}'
            txt_file.write(content + '\n')

            # Update verification item and quantity if they exist
            if verification_item_tmp is not None and verification_quantity_tmp is not None:
                # content = f'核銷報單項次:{verification_item} 核銷數量:{verification_quantity}'
                # txt_file.write(content + '\n')
                verification_item = verification_item_tmp
                verification_quantity = verification_quantity_tmp

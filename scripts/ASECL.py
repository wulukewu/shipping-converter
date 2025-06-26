import openpyxl
from openpyxl.utils import get_column_letter
import xlrd
import os

def convert_xls_to_xlsx(xls_filename, xlsx_filename):
    """
    Converts an XLS file to XLSX format.
    Args:
        xls_filename (str): The path to the XLS file.
        xlsx_filename (str): The path where the XLSX file will be saved.
    """
    try:
        workbook = xlrd.open_workbook(xls_filename)
        wb = openpyxl.Workbook()
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            new_sheet = wb.create_sheet(title=sheet_name)
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    new_sheet.cell(row=row + 1, column=col + 1, value=cell_value)
        # Remove the default sheet if it's still present
        if "Sheet" in wb.sheetnames:
          default_sheet = wb["Sheet"]
          wb.remove(default_sheet)
        wb.save(xlsx_filename)
        print(f"Successfully converted '{xls_filename}' to '{xlsx_filename}'")
    except Exception as e:
        print(f"Error converting XLS to XLSX: {e}")
        return False
    return True

def organize_data(filename):
    """
    Organizes data from '實際進出倉明細' sheets in an Excel file
    and writes the organized data to a new Excel file.
    Args:
        filename (str): The path to the Excel file containing '實際進出倉明細' sheets.
    """
    # Load the workbook with data_only=True to get only values
    wb = openpyxl.load_workbook(filename, data_only=True)

    # Get the main sheet ('實際進出倉明細 ')
    if "實際進出倉明細 " not in wb.sheetnames:
        print("Sheet named '實際進出倉明細 ' not found. Please make sure it exists.")
        return
    main_sheet = wb["實際進出倉明細 "]

    # Create a new workbook for the organized sheet
    organized_wb = openpyxl.Workbook()
    organized_sheet = organized_wb.active
    organized_sheet.title = "Organized"

    # Initialize item number counter for "Organized" sheet
    j = 1  # Initialize j for the "Organized" sheet row index

    last_row = main_sheet.max_row

    i = 2  # Initialize i for the main sheet row index
    
    while i <= last_row:
        item_idx = main_sheet.cell(row=i, column=1).value
        item_des = main_sheet.cell(row=i, column=2).value
        item_part_no = main_sheet.cell(row=i, column=3).value
        item_unit = main_sheet.cell(row=i, column=4).value
        item_currency = main_sheet.cell(row=i, column=5).value
        item_po_no = main_sheet.cell(row=i, column=6).value
        item_org_no = main_sheet.cell(row=i, column=7).value
        item_org_idx = main_sheet.cell(row=i, column=8).value
        item_name = main_sheet.cell(row=i, column=9).value
        item_tariff = main_sheet.cell(row=i, column=10).value
        item_origin = main_sheet.cell(row=i, column=11).value

        # Break the loop if we reach an empty row (no item index)
        if item_idx is None:
            break

        # Write data to the organized sheet with specific formatting
        organized_sheet.cell(row=j, column=1).value = item_des      # Item Description
        organized_sheet.cell(row=j, column=2).value = item_part_no  # Part Number
        organized_sheet.cell(row=j, column=3).value = item_unit     # Unit
        organized_sheet.cell(row=j, column=4).value = ''            # Empty column
        organized_sheet.cell(row=j, column=5).value = item_des      # Item Description (duplicate)
        organized_sheet.cell(row=j, column=6).value = 'NO BRAND'    # Brand information
        organized_sheet.cell(row=j, column=7).value = item_tariff   # Tariff code
        organized_sheet.cell(row=j, column=8).value = '06'          # Fixed value '06'
        organized_sheet.cell(row=j, column=9).value = item_org_no   # Organization number
        organized_sheet.cell(row=j, column=10).value = item_org_idx # Organization index
        organized_sheet.cell(row=j, column=11).value = 'YB'         # Fixed value 'YB'

        i += 1  # Move to next row in source sheet
        j += 1  # Move to next row in organized sheet

    # Autofit all rows in organized sheet for better readability
    for row in organized_sheet.rows:
        for cell in row:
            if cell.value:
                organized_sheet.row_dimensions[cell.row].auto_size = True

    # Save the organized data to a new file in the 'uploads' folder
    output_filename = os.path.join("uploads", "Organized_Data.xlsx")
    organized_wb.save(output_filename)
    print(f"Organized data saved to '{output_filename}'")
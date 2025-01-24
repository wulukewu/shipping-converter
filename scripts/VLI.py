import openpyxl
from openpyxl.utils import get_column_letter
import xlrd
import os

def convert_xls_to_xlsx(xls_filename, xlsx_filename):
    """
    Converts an XLS file to XLSX format.
    Args:
        xls_filename (str): The path to the XLS file.
        xlsx_filename (str): The path where the XLSX file will be saved.
    """
    try:
        workbook = xlrd.open_workbook(xls_filename)
        wb = openpyxl.Workbook()
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            new_sheet = wb.create_sheet(title=sheet_name)
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    new_sheet.cell(row=row + 1, column=col + 1, value=cell_value)
        # Remove the default sheet if it's still present
        if "Sheet" in wb.sheetnames:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)
        wb.save(xlsx_filename)
        print(f"Successfully converted '{xls_filename}' to '{xlsx_filename}'")
    except Exception as e:
        print(f"Error converting XLS to XLSX: {e}")
        return False
    return True

def organize_data(filename):
    """
    Organizes data from the main sheet in an Excel file
    and writes the organized data to a new Excel file.

    Args:
    filename (str): The path to the Excel file containing the main sheet.
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(filename)

        # Get the (first) main sheet
        main_sheet = wb.worksheets[0]

        # Create a new workbook for the organized sheet
        organized_wb = openpyxl.Workbook()
        organized_sheet = organized_wb.active
        organized_sheet.title = "Organized"

        # Initialize row indices for the main and organized sheets
        i = 1  # Row index for the main sheet
        j = 1  # Row index for the organized sheet

        # Find the last row in the main sheet
        last_row = main_sheet.max_row

        while i <= last_row:
            part_no = main_sheet.cell(row=i, column=1).value
            if part_no is None:
                break

            # Read data from the main sheet
            desc_1 = main_sheet.cell(row=i, column=2).value
            qty = int(main_sheet.cell(row=i, column=3).value)
            unit_price = float(main_sheet.cell(row=i, column=4).value)
            amount = main_sheet.cell(row=i, column=5).value
            po_no = main_sheet.cell(row=i+1, column=1).value
            desc_2 = main_sheet.cell(row=i+1, column=2).value
            qty_unit = main_sheet.cell(row=i+1, column=3).value
            unit_price_unit = main_sheet.cell(row=i+1, column=4).value
            amount_unit = main_sheet.cell(row=i+1, column=5).value

            # Write data to the organized sheet
            organized_sheet.cell(row=j, column=1, value=f'Part No.{part_no}')
            organized_sheet.cell(row=j, column=2, value=f'Po. No.{po_no}')
            organized_sheet.cell(row=j, column=3, value=desc_1)
            organized_sheet.cell(row=j, column=4, value=desc_2)
            organized_sheet.cell(row=j, column=5, value=qty)
            organized_sheet.cell(row=j, column=6, value=qty_unit)
            organized_sheet.cell(row=j, column=7, value=unit_price)
            organized_sheet.cell(row=j, column=8, value='VLI')
            organized_sheet.cell(row=j, column=9, value='85423900228')
            organized_sheet.cell(row=j, column=10, value='02')

            # Format quantity cell
            qty_cell = organized_sheet.cell(row=j, column=5, value=qty)
            if isinstance(qty, (int, float)):
                qty_cell.number_format = '#,##0'
            
            # Format unit price cell
            unit_price_cell = organized_sheet.cell(row=j, column=7, value=unit_price)
            if isinstance(unit_price, (int, float)):
                unit_price_cell.number_format = '#,##0.0000'

            # Increment the row indices
            i += 3
            j += 1

        # Adjust column widths
        for col in range(1, 10):
            max_length = 0
            column = get_column_letter(col)
            for cell in organized_sheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            organized_sheet.column_dimensions[column].width = adjusted_width

        # Save the organized data to a new file in the 'uploads' folder
        output_filename = os.path.join("uploads", "Organized_Data.xlsx")
        organized_wb.save(output_filename)
        print(f"Organized data saved to '{output_filename}'")

    except FileNotFoundError:
        print(f"Error: File not found at '{filename}'.")
    except Exception as e:
        print(f"An error occurred: {e}")
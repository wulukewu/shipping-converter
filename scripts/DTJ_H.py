import openpyxl
from openpyxl.utils import get_column_letter
import re
import xlrd
import os

def convert_xls_to_xlsx(xls_filename, xlsx_filename):
    """
    Converts an XLS file to XLSX format.
    Args:
        xls_filename (str): The path to the XLS file.
        xlsx_filename (str): The path where the XLSX file will be saved.
    Returns:
        bool: True if the conversion was successful, False otherwise.
    """
    try:
        # Open the XLS workbook
        workbook = xlrd.open_workbook(xls_filename)
        # Create a new XLSX workbook
        wb = openpyxl.Workbook()

        # Iterate through each sheet in the XLS workbook
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            new_sheet = wb.create_sheet(title=sheet_name)

            # Copy data from each cell in the XLS sheet to the XLSX sheet
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    new_sheet.cell(row=row + 1, column=col + 1, value=cell_value)

        # Remove the default sheet if it exists (named "Sheet")
        if "Sheet" in wb.sheetnames:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)

        # Save the XLSX workbook
        wb.save(xlsx_filename)
        print(f"Successfully converted '{xls_filename}' to '{xlsx_filename}'")
        return True

    except Exception as e:
        print(f"Error converting XLS to XLSX: {e}")
        return False


def organize_data(filename):
    """
    Organizes data from the 'HM' sheet in an Excel file and writes the organized data to a new Excel file
    named "Organized_Data.xlsx" in the 'uploads' folder.  Specifically extracts description of goods,
    unit price, quantity, and kind, processes the 'free sample' entries and general entries.
    Args:
        filename (str): The path to the Excel file containing the 'HM' sheet.
    """
    # Load the workbook
    wb = openpyxl.load_workbook(filename)

    # Get the HM sheet
    if "HM" not in wb.sheetnames:
        print("Sheet named 'HM' not found. Please make sure it exists.")
        return

    hm_sheet = wb["HM"]

    # Create a new workbook for the organized sheet
    organized_wb = openpyxl.Workbook()
    org_sheet = organized_wb.active  # Get the active sheet
    org_sheet.title = "Organized"
    ctn_sheet = organized_wb.create_sheet(title="ctn")

    # Find the last row of the HM sheet
    last_row = hm_sheet.max_row

    # Find the start and end rows for data extraction based on markers in column A
    found_start = False
    found_end = False
    start_row = 0
    end_row = 0

    # Loop through column A to find start and end points
    for i in range(1, last_row + 1):
        cell_value = hm_sheet.cell(row=i, column=1).value
        if not found_start:
            if cell_value and "DESCRIPTION OF GOODS" in str(cell_value).upper():
                start_row = i + 1
                found_start = True
        elif not found_end:
            if cell_value and "PACKING AND WEIGHT LIST" in str(cell_value).upper():
                end_row = i - 1
                found_end = True
                break

    # Check if we found both the start and end markers
    if not found_start or not found_end:
        print("Could not find start or end markers")
        return

    # Initialize row counter for the organized sheet
    j = 1
    # Initialize list to store 'kind' values
    kinds = []
    # Initialize minimum unit price to a very high value
    min_unit_price = float('inf')

    # Skip the first row after the start marker
    start_row += 1

    # Loop through the rows between the start and end rows in HM sheet
    for i in range(start_row, end_row + 1):
        # Get data from goods
        desc = hm_sheet.cell(row=i, column=1).value
        unit_price = hm_sheet.cell(row=i, column=6).value
        qty = hm_sheet.cell(row=i, column=7).value
        kind = hm_sheet.cell(row=i, column=10).value

        # Skip to the next iteration if description is blank
        if not desc or str(desc).strip() == "":
            continue

        # Handle 'free sample' entries
        if desc and "free sample" in str(desc).lower():
            # Case 2: 'free sample' set
            kinds_combined = ""  # Store the combined 'kind' values
            # Ensure that the min price is not infinite
            if min_unit_price == float('inf'):
                min_unit_price = 0
            sub_kind_count = 0

            if len(kinds) > 1:
                # Case 2-1:  'kinds' list has multiple entries
                for k in range(len(kinds)):
                    # Check for "xN" multiplier in sub-kind
                    sub_kind_without_multiplier = kinds[k]
                    multiplier = 1  # default value for multiplier
                    if "x" in str(kinds[k]).lower():
                        # Extract multiplier value
                        parts = re.split(r"x", kinds[k], flags=re.IGNORECASE)  # Split case-insensitive
                        if len(parts) == 2 and parts[1].isdigit():
                            multiplier = int(parts[1])
                            sub_kind_without_multiplier = kinds[k]
                    sub_kind_count += multiplier
                    kinds_combined += f"({sub_kind_without_multiplier})-{qty}PCS" + (
                        f"x{multiplier}" if multiplier > 1 else "") + "\n"
            else:
                # Case 2-2:  'kinds' list has one entry or is empty
                if len(kinds) > 0:
                    for k in range(len(kinds)):
                        # Check for "xN" multiplier in sub-kind
                        sub_kind_without_multiplier = kinds[k]
                        multiplier = 1  # default value for multiplier
                        if "x" in str(kinds[k]).lower():
                            # Extract multiplier value
                            parts = re.split(r"x", kinds[k], flags=re.IGNORECASE)  # Split case-insensitive
                            if len(parts) == 2 and parts[1].isdigit():
                                multiplier = int(parts[1])
                                sub_kind_without_multiplier = kinds[k]
                        sub_kind_count += multiplier
                        kinds_combined = f"({sub_kind_without_multiplier})"

            if len(kinds_combined) > 2 and len(kinds) > 1:
                kinds_combined = kinds_combined[:-1]  # Remove the trailing newline character

            org_sheet.cell(row=j, column=2).value = kinds_combined

            if len(kinds) > 1:
                # Case 2-1
                desc_parts = re.split(r"free sample", desc, flags=re.IGNORECASE)
                org_sheet.cell(row=j, column=1).value = f"{desc_parts[0].strip()} ({len(kinds)}pcs/set)\nfree sample"
                org_sheet.cell(row=j, column=1).alignment = openpyxl.styles.Alignment(wrap_text=True)
                org_sheet.cell(row=j, column=4).value = "SET"
            else:
                # Case 2-2
                org_sheet.cell(row=j, column=1).value = desc
                org_sheet.cell(row=j, column=4).value = "PCS"

            org_sheet.cell(row=j, column=3).value = qty
            org_sheet.cell(row=j, column=5).value = min_unit_price
            org_sheet.cell(row=j, column=6).number_format = "@" # Ensure text format
            org_sheet.cell(row=j, column=6).value = "04"
            org_sheet.cell(row=j, column=7).value = "F.O.C."  # Free of Charge
            org_sheet.cell(row=j, column=8).value = "NO BRAND"
            org_sheet.cell(row=j, column=9).number_format = "@" # Ensure text format
            org_sheet.cell(row=j, column=9).value = "49119900002"

            # Clear the 'kinds' list for the next iteration
            kinds = []
            min_unit_price = float('inf')  # Reset min_unit_price
        else:
            # Case 1:  Normal entries (not 'free sample')
            # Add the 'kind' value to the 'kinds' list

            # Remove parentheses from the kind string
            if kind and kind.startswith("(") and kind.endswith(")"):
                kind = kind[1:-1]

            # Split the kind by commas into the subkinds array
            sub_kinds = []
            if kind:
                sub_kinds = kind.split(",")

            sub_kinds_combined = ""
            qty_string = f"{qty:,}PCS"
            sub_kind_count = 0
            for sub_k in sub_kinds:
                # Check for "xN" multiplier in sub-kind
                sub_kind_without_multiplier = sub_k.strip()
                multiplier = 1  # default value for multiplier
                if "x" in str(sub_k).lower():
                    # Extract multiplier value
                    parts = re.split(r"x", sub_k, flags=re.IGNORECASE)  # Split case-insensitive
                    if len(parts) == 2 and parts[1].isdigit():
                        multiplier = int(parts[1])
                        sub_kind_without_multiplier = sub_k.strip()
                sub_kind_count += multiplier
                if len(sub_kinds) > 1:
                    sub_kinds_combined += f"({sub_kind_without_multiplier})-{qty_string}" + (
                        f"x{multiplier}" if multiplier > 1 else "") + "\n"
                else:
                    sub_kinds_combined = f"({sub_kind_without_multiplier})"
                kinds.append(sub_kind_without_multiplier)

            if len(sub_kinds_combined) > 2 and len(sub_kinds) > 1:
                sub_kinds_combined = sub_kinds_combined[:-1]

            # Construct the cell A value with the (n pcs/set) part if needed
            if sub_kind_count > 1:
                org_sheet.cell(row=j, column=1).value = f"{desc} ({sub_kind_count}pcs/set)"
                org_sheet.cell(row=j, column=4).value = "SET"
            else:
                org_sheet.cell(row=j, column=1).value = desc
                org_sheet.cell(row=j, column=4).value = "PCS"

            org_sheet.cell(row=j, column=2).value = sub_kinds_combined
            org_sheet.cell(row=j, column=3).value = qty
            org_sheet.cell(row=j, column=5).value = unit_price

            # Update the min price
            if unit_price < min_unit_price:
                min_unit_price = unit_price

            org_sheet.cell(row=j, column=6).number_format = "@" # Ensure text format
            org_sheet.cell(row=j, column=6).value = "02"
            org_sheet.cell(row=j, column=8).value = "NO BRAND"
            org_sheet.cell(row=j, column=9).number_format = "@" # Ensure text format
            org_sheet.cell(row=j, column=9).value = "49119900002"

        # Increment the row counter for orgSheet
        j += 1

    # Format the "Organized" sheet columns
    org_sheet.column_dimensions[get_column_letter(1)].width = 32.33
    org_sheet.column_dimensions[get_column_letter(2)].width = 30
    org_sheet.column_dimensions[get_column_letter(3)].number_format = "#,##0"

    # Autofit all rows and columns in organized sheet
    for row in org_sheet.rows:
        for cell in row:
            if cell.value:
                org_sheet.row_dimensions[cell.row].auto_size = True
    for column_cells in org_sheet.columns:
        org_sheet.column_dimensions[column_cells[0].column_letter].auto_size = True

    # =====================================
    # CTN SHEET IMPLEMENTATION STARTS HERE
    # =====================================
    
    def normalize_format(text):
        """
        Normalizes the format by adding a space before parentheses if it's missing.
        Example: "TCED-8038(dvd jacket)" becomes "TCED-8038 (dvd jacket)"
        """
        if not text:
            return text
        
        # Use regex to add space before opening parenthesis if it's missing
        # This matches a pattern like "TCED-8038(dvd jacket)" and adds space before "("
        normalized = re.sub(r'([A-Z0-9-])\s*\(', r'\1 (', str(text))
        return normalized
    
    def normalize_ctn_format(text):
        """
        Normalizes the 'ctn no.HM' format by ensuring there's a space between HM and the number.
        Example: "ctn no.HM25" becomes "ctn no.HM 25"
        """
        if not text:
            return text
        
        # Use regex to add space between HM and number if it's missing
        normalized = re.sub(r'(ctn no\.HM)(\d+)', r'\1 \2', str(text), flags=re.IGNORECASE)
        return normalized
    
    found_start_ctn = False
    found_end_ctn = False
    start_row_ctn = 0
    end_row_ctn = 0
    j_ctn = 0

    # Loop through column A to find start and end points
    for i in range(1, last_row + 1):
        cell_value = hm_sheet.cell(row=i, column=1).value
        if j_ctn < 1 and cell_value and "CD PRINTING MATTERS" in str(cell_value).upper():
            j_ctn = i + 1
        if not found_start_ctn:
            if cell_value and "QTY FOR EACH CTN" in str(cell_value).upper():
                start_row_ctn = i + 1
                found_start_ctn = True
        elif not found_end_ctn:
            if cell_value and "pallet".upper() in str(cell_value).upper():
                end_row_ctn = i
                found_end_ctn = True
                break

    # Check if we found both the start and end markers
    if not found_start_ctn or not found_end_ctn:
        print("Could not find start or end markers for CTN sheet")
        return

    # Initialize k for the organized sheet row
    k = 1

    # Skip the first row from start
    start_row_ctn += 1

    desc_list = []
    goods = None
    initial_j_ctn = j_ctn  # Store the initial position

    # Loop through the rows between the start and end rows in HM sheet
    for i in range(start_row_ctn, end_row_ctn + 1):
        ctn = hm_sheet.cell(row=i, column=1).value
        qty = hm_sheet.cell(row=i, column=4).value
        net_weight = hm_sheet.cell(row=i, column=6).value
        gross_weight = hm_sheet.cell(row=i, column=9).value

        # Check if ctn is blank, if it is, skip to the next iteration
        if not ctn or str(ctn).strip() == "":
            continue

        # Normalize the ctn format
        ctn = normalize_format(ctn)

        # print(f'CTN: {ctn}, QTY: {qty}, Net Weight: {net_weight}, Gross Weight: {gross_weight}')

        if qty is None or 'pallet' in str(ctn).lower():
            goods_tmp = normalize_format(hm_sheet.cell(row=i + 1, column=1).value).split(' ')[0]
            if goods is None:
                goods = goods_tmp
            elif goods != goods_tmp:
                pcs_per_set = 0
                j_ctn = initial_j_ctn  # Reset to initial position for new goods
                
                # First pass: count the number of items for this goods
                temp_j_ctn = initial_j_ctn
                sub_kind_count = 0
                while True:
                    desc_goods = normalize_format(str(hm_sheet.cell(row=temp_j_ctn, column=1).value))
                    if 'free sample' in desc_goods and goods in desc_goods:
                        break
                    elif goods in desc_goods:
                        # Get the kind value from column 10
                        kind = hm_sheet.cell(row=temp_j_ctn, column=10).value
                        if kind and kind.startswith("(") and kind.endswith(")"):
                            kind = kind[1:-1]
                        
                        # Split the kind by commas
                        sub_kinds = []
                        if kind:
                            sub_kinds = kind.split(",")
                        
                        # Count the items in this kind
                        for sub_k in sub_kinds:
                            sub_kind_without_multiplier = sub_k.strip()
                            multiplier = 1
                            if "x" in str(sub_k).lower():
                                parts = re.split(r"x", sub_k, flags=re.IGNORECASE)
                                if len(parts) == 2 and parts[1].isdigit():
                                    multiplier = int(parts[1])
                                    sub_kind_without_multiplier = sub_k.strip()
                            sub_kind_count += multiplier
                        temp_j_ctn += 1
                    else:
                        temp_j_ctn += 1
                
                # Second pass: find the free sample and use its quantity
                while True:
                    desc_goods = normalize_format(str(hm_sheet.cell(row=j_ctn, column=1).value))
                    # print(f'Goods: {goods}, Desc: {desc_goods}')
                    if 'free sample' in desc_goods and goods in desc_goods:
                        # Use the original quantity from free sample row
                        free_sample_qty = hm_sheet.cell(row=j_ctn, column=7).value
                        
                        # Determine unit based on sub_kind_count
                        if sub_kind_count > 1:
                            free_sample_qty_desc = 'SET'
                        else:
                            free_sample_qty_desc = 'PCS'
                        
                        ctn_sheet.cell(row=k, column=1).value = f'{goods}\n({free_sample_qty}{free_sample_qty_desc}-free sample)'
                        j_ctn += 1
                        k += 1
                        break
                    elif goods in desc_goods:
                        pcs_per_set += 1
                        j_ctn += 1
                    else:
                        j_ctn += 1

                ctn_no_list = []
                for desc in desc_list:
                    # Use regex to match 'ctn no.HM' with optional space before numbers
                    ctn_match = re.search(r'ctn no\.HM\s*(\d+(?:-\d+)?)', desc, re.IGNORECASE)
                    if ctn_match:
                        number_part = ctn_match.group(1)
                        if '-' in number_part:
                            ctn_no_list.extend(list(map(int, number_part.split('-'))))
                        else:
                            ctn_no_list.append(int(number_part))
                
                if len(ctn_no_list) > 1:
                    desc_list = [desc if not re.search(r'ctn no\.HM\s*\d+', desc, re.IGNORECASE) else f"\n{desc}" for desc in desc_list]
                    for desc_idx in range(len(desc_list)):
                        if '\n' in desc_list[desc_idx] and re.search(r'ctn no\.HM\s*\d+', desc_list[desc_idx], re.IGNORECASE):
                            desc_list[desc_idx] = desc_list[desc_idx].replace('\n', '')
                            break

                    ctn_no_num = max(ctn_no_list) - min(ctn_no_list) + 1
                    desc_list.append(f'x {ctn_no_num}CTNS')
                else:
                    desc_list = [desc for desc in desc_list if not re.search(r'ctn no\.HM\s*\d+', desc, re.IGNORECASE)]

                desc = '\n'.join(desc_list).strip()
                ctn_sheet.cell(row=k, column=1).value = desc
                k += 1
                desc_list = []
                goods = goods_tmp
            if re.search(r'ctn no\.HM\s*\d+', str(ctn), re.IGNORECASE):
                desc_list.append(normalize_ctn_format(ctn))
        else:
            desc_list.append(f'{ctn}-{qty}PCS')

    # Autofit all rows and columns in ctn sheet
    for row in ctn_sheet.rows:
        for cell in row:
            if cell.value:
                ctn_sheet.row_dimensions[cell.row].auto_size = True
    for column_cells in ctn_sheet.columns:
        ctn_sheet.column_dimensions[column_cells[0].column_letter].auto_size = True

    # Save the organized data to a new file in the 'uploads' folder
    output_filename = os.path.join("uploads", "Organized_Data.xlsx")
    organized_wb.save(output_filename)
    print(f"Organized data saved to '{output_filename}'")
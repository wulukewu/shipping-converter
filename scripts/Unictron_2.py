import openpyxl
from openpyxl.utils import get_column_letter
import xlrd
import os
import json

def convert_xls_to_xlsx(xls_filename, xlsx_filename):
    """
    Converts an XLS file to XLSX format.
    Args:
        xls_filename (str): The path to the XLS file.
        xlsx_filename (str): The path where the XLSX file will be saved.
    """
    try:
        workbook = xlrd.open_workbook(xls_filename)
        wb = openpyxl.Workbook()
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            new_sheet = wb.create_sheet(title=sheet_name)
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    new_sheet.cell(row=row + 1, column=col + 1, value=cell_value)
        # Remove the default sheet if it's still present
        if "Sheet" in wb.sheetnames:
          default_sheet = wb["Sheet"]
          wb.remove(default_sheet)
        wb.save(xlsx_filename)
        print(f"Successfully converted '{xls_filename}' to '{xlsx_filename}'")
    except Exception as e:
        print(f"Error converting XLS to XLSX: {e}")
        return False
    return True

def load_hs_codes():
    """
    Load HS codes from JSON file.
    
    Returns:
    dict: Dictionary containing HS codes mapping
    """
    try:
        json_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'Unictron_2_hs_codes.json')
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: Could not load HS codes JSON file: {e}")
        return {}

def get_hs_code_from_desc(item_desc, hs_codes_dict):
    """
    Extract HS code from item description.
    
    Args:
    item_desc (str): Item description
    hs_codes_dict (dict): Dictionary of HS codes
    
    Returns:
    str: HS code or "NA" if not found
    """
    if not item_desc:
        return "NA"
    
    # Split the description and check all words
    desc_parts = str(item_desc).split()
    if desc_parts:
        for part in desc_parts:
            part_upper = part.upper()
            if part_upper in hs_codes_dict:
                return hs_codes_dict[part_upper]["hs_code"]
    
    return "NA"

def organize_data(filename):
    """
    Organizes data from 'Invoice' and 'Booking' sheets in an Excel file
    and writes the organized data to a new Excel file.

    Args:
    filename (str): The path to the Excel file containing 'Invoice' and 'Booking' sheets.
    """
    # Load the workbook
    wb = openpyxl.load_workbook(filename)

    # Get the main sheet ('Invoice')
    if "Invoice" not in wb.sheetnames:
        print("Sheet named 'Invoice' not found. Please make sure it exists.")
        return
    main_sheet = wb["Invoice"]

    # Load HS codes dictionary
    hs_codes_dict = load_hs_codes()
    print(f"Loaded {len(hs_codes_dict)} HS codes from JSON file")
    
    # Create a new workbook for the organized sheet
    organized_wb = openpyxl.Workbook()
    organized_sheet = organized_wb.active
    organized_sheet.title = "Organized"

    # Initialize item number counter for "Organized" sheet
    item_number = 1
    j = 1  # Initialize j for the "Organized" sheet row index

    # Find the last row in column B of the main sheet
    last_row = main_sheet.max_row

    i = 1  # Initialize i for the main sheet row index
    start_row = -1
    
    while i <= last_row:
        cell_value = main_sheet.cell(row=i, column=1).value
        if cell_value and "序号" in str(cell_value):
            i += 1
            start_row = i
            continue
        elif start_row == -1:
            i += 1
            continue
        elif "TOTAL" in str(cell_value):
            break
        elif cell_value is None:
            break

        item_number = main_sheet.cell(row=i, column=1).value
        item_material = main_sheet.cell(row=i, column=2).value
        item_vendor_number = main_sheet.cell(row=i, column=3).value
        item_po_number = main_sheet.cell(row=i, column=4).value
        item_item = main_sheet.cell(row=i, column=5).value
        item_qty = int(main_sheet.cell(row=i, column=6).value)
        item_unit = main_sheet.cell(row=i, column=7).value
        item_amount = main_sheet.cell(row=i, column=8).value
        item_country = main_sheet.cell(row=i, column=9).value
        item_desc = main_sheet.cell(row=i, column=10).value

        print(f'Item Number: {item_number}, Material: {item_material}, Vendor Number: {item_vendor_number}, PO Number: {item_po_number}, Item: {item_item}, QTY: {item_qty}, Unit: {item_unit}, Amount: {item_amount}, Country: {item_country}, Desc: {item_desc}')

        # Get HS code based on item description
        hs_code = get_hs_code_from_desc(item_desc, hs_codes_dict)
        
        organized_sheet.cell(row=j, column=1).value = item_desc
        organized_sheet.cell(row=j, column=2).value = item_material
        organized_sheet.cell(row=j, column=3).value = item_vendor_number
        organized_sheet.cell(row=j, column=4).value = item_po_number
        
        # Set quantity with number formatting
        qty_cell = organized_sheet.cell(row=j, column=5)
        qty_cell.value = item_qty
        qty_cell.number_format = "#,##0"
        
        organized_sheet.cell(row=j, column=6).value = "PCS"
        organized_sheet.cell(row=j, column=7).value = item_unit
        organized_sheet.cell(row=j, column=8).value = "NO BRAND"
        organized_sheet.cell(row=j, column=9).value = hs_code
        organized_sheet.cell(row=j, column=10).value = "02"
        
        i += 1
        j += 1


    # Auto-adjust width for all columns
    for col in range(1, organized_sheet.max_column + 1):
        column_letter = get_column_letter(col)
        # Calculate the maximum width needed for this column
        max_width = 0
        for row in range(1, organized_sheet.max_row + 1):
            cell = organized_sheet.cell(row=row, column=col)
            if cell.value:
                # Estimate width based on content length
                content_length = len(str(cell.value))
                # Add some padding
                estimated_width = content_length + 2
                max_width = max(max_width, estimated_width)
        
        # Set the column width (minimum 8, maximum 50)
        column_width = max(8, min(max_width, 50))
        organized_sheet.column_dimensions[column_letter].width = column_width

    # Autofit all rows in organized sheet
    for row in organized_sheet.rows:
       for cell in row:
            if cell.value:
                 organized_sheet.row_dimensions[cell.row].auto_size = True

    # Save the organized data to a new file in the 'uploads' folder
    output_filename = os.path.join("uploads", "Organized_Data.xlsx")
    organized_wb.save(output_filename)
    print(f"Organized data saved to '{output_filename}'")
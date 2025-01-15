import openpyxl
from openpyxl.utils import get_column_letter
import os

def organize_data(filename):
    """
    Organizes data from 'Invoice' and 'Booking' sheets in an Excel file
    and writes the organized data to a new Excel file.

    Args:
    filename (str): The path to the Excel file containing 'Invoice' and 'Booking' sheets.
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(filename)

        # Get the main sheet ('Invoice')
        if "Invoice" not in wb.sheetnames:
            print("Sheet named 'Invoice' not found. Please make sure it exists.")
            return
        main_sheet = wb["Invoice"]

        # Get the booking sheet ('Booking')
        if "Booking" not in wb.sheetnames:
            print("Sheet named 'Booking' not found. Please make sure it exists.")
            return
        booking_sheet = wb["Booking"]

        # Find "description" and "HS CODE" columns in booking sheet
        des_col = 0
        hs_col = 0
        last_col = booking_sheet.max_column
        for i in range(1, last_col + 1):
            cell_value = booking_sheet.cell(row=1, column=i).value
            if cell_value and cell_value.strip().lower() == "description":
                des_col = i
            elif cell_value and cell_value.strip().lower() == "hs code":
                hs_col = i
            if des_col > 0 and hs_col > 0:
                break
        if des_col == 0:
            print("Description column was not found. Code execution stopped.")
            return
        if hs_col == 0:
            print("HS CODE column was not found. Code execution stopped.")
            return

        # Create a new workbook for the organized sheet
        organized_wb = openpyxl.Workbook()
        organized_sheet = organized_wb.active
        organized_sheet.title = "Organized"

        # Initialize item number counter for "Organized" sheet
        item_number = 1
        j = 1  # Initialize j for the "Organized" sheet row index

        # Find the last row in column B of the main sheet
        last_row = main_sheet.max_row
        booking_last_row = booking_sheet.max_row

        i = 1  # Initialize i for the main sheet row index
        
        while i <= last_row:
            item_des = main_sheet.cell(row=i+1, column=2).value # Get the next Item Description before checking part no or po no
            if main_sheet.cell(row=i, column=2).value == "PART NO:":
                # Case 1: "PART NO:" found
                part_info = f"{main_sheet.cell(row=i, column=2).value} {main_sheet.cell(row=i, column=3).value}\n{main_sheet.cell(row=i + 1, column=2).value}"
                organized_sheet.cell(row=j, column=1).value = part_info
                i += 2
                j += 1
            elif main_sheet.cell(row=i, column=2).value == "P.O.NO: ":
                # Case 2: "P.O.NO:" found
                po_info = f"{main_sheet.cell(row=i, column=2).value} {main_sheet.cell(row=i, column=3).value}"
                organized_sheet.cell(row=j, column=1).value = po_info
                organized_sheet.cell(row=j, column=2).value = main_sheet.cell(row=i, column=4).value
                organized_sheet.cell(row=j, column=3).value = "PCS"
                organized_sheet.cell(row=j, column=4).value = main_sheet.cell(row=i, column=5).value
                organized_sheet.cell(row=j, column=5).value = "NO BRAND"
                organized_sheet.cell(row=j, column=7).number_format = "@"
                organized_sheet.cell(row=j, column=7).value = "02"
               # Initialize the found flag
                found = False
                # Loop through each row in booking sheet
                for k in range(1, booking_last_row):
                    # Check if the item description matches
                    booking_des = booking_sheet.cell(row=k, column=des_col).value
                    if booking_des and item_des and booking_des.strip().lower() == item_des.strip().lower():
                    # Item description found
                        organized_sheet.cell(row=j, column=6).value = booking_sheet.cell(row=k, column=hs_col).value
                        found = True  # Set the flag to indicate that the item was found
                        break      # Exit the loop since the item was found
                if not found:
                    organized_sheet.cell(row=j, column=6).value = "N/A"
                i += 1
                j += 1
            else:
                # Case 3: Skip the row
                i += 1

        # Format the "Organized" sheet columns
        organized_sheet.column_dimensions[get_column_letter(1)].width = 31.83
        organized_sheet.column_dimensions[get_column_letter(2)].width = 8.33
        organized_sheet.column_dimensions[get_column_letter(2)].number_format = "#,##0"
        organized_sheet.column_dimensions[get_column_letter(3)].auto_size = True
        organized_sheet.column_dimensions[get_column_letter(4)].width = 7
        organized_sheet.column_dimensions[get_column_letter(4)].number_format = "0.00"
        organized_sheet.column_dimensions[get_column_letter(5)].auto_size = True
        organized_sheet.column_dimensions[get_column_letter(6)].auto_size = True
        organized_sheet.column_dimensions[get_column_letter(7)].auto_size = True

         #Autofit all rows in organized sheet
        for row in organized_sheet.rows:
           for cell in row:
                if cell.value:
                     organized_sheet.row_dimensions[cell.row].auto_size = True

        # Save the organized data to a new file in the 'uploads' folder
        output_filename = os.path.join("uploads", "Organized_Data.xlsx")
        organized_wb.save(output_filename)
        print(f"Organized data saved to '{output_filename}'")

    except FileNotFoundError:
        print(f"Error: File not found at '{filename}'.")
    except Exception as e:
         print(f"An error occurred: {e}")